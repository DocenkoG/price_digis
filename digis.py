# -*- coding: UTF-8 -*-
import os
import os.path
import logging
import logging.config
import sys
import configparser
import time
import shutil
import openpyxl                       # Для .xlsx
#import xlrd                          # для .xls
from   price_tools import getCellXlsx, getCell, quoted, dump_cell, currencyType, openX, sheetByName
import csv


def getXlsxString(sh, i, in_columns_j):
    impValues = {}
    for item in in_columns_j.keys():
        j = in_columns_j[item]
        if item in ('закупка', 'продажа', 'цена'):
            if getCellXlsx(row=i, col=j, isDigit='N', sheet=sh).find('звоните') >= 0:
                impValues[item] = '0.1'
            else:
                impValues[item] = getCellXlsx(row=i, col=j, isDigit='Y', sheet=sh)
            #print(sh, i, sh.cell( row=i, column=j).value, sh.cell(row=i, column=j).number_format, currencyType(sh, i, j))
        elif item == 'валюта_по_формату':
            impValues[item] = currencyType(row=i, col=j, sheet=sh)
        else:
            impValues[item] = getCellXlsx(row=i, col=j, isDigit='N', sheet=sh)
    return impValues



def convert_excel2csv(cfg):
    csvFNameRUR  = cfg.get('basic','filename_out_RUR')
    csvFNameEUR  = cfg.get('basic','filename_out_EUR')
    csvFNameUSD  = cfg.get('basic','filename_out_USD')
    csvFNameRUR1 = 'csv_digis_RUR1.csv'
    csvFNameEUR1 = 'csv_digis_EUR1.csv'
    csvFNameUSD1 = 'csv_digis_USD1.csv'
    csvFNameRUR2 = 'csv_digis_RUR2.csv'
    csvFNameEUR2 = 'csv_digis_EUR2.csv'
    csvFNameUSD2 = 'csv_digis_USD2.csv'

    priceFName= cfg.get('basic','filename_in')
    sheetName = cfg.get('basic','sheetname')
    
    log.debug('Reading file ' + priceFName )
    sheet = sheetByName(fileName = priceFName, sheetName = sheetName)
    if not sheet :
        log.error("Нет листа "+sheetName+" в файле "+ priceFName)
        return False
    log.debug("Sheet   "+sheetName)
    out_cols = cfg.options("cols_out")
    in_cols  = cfg.options("cols_in")
    out_template = {}
    for vName in out_cols :
         out_template[vName] = cfg.get("cols_out", vName)
    in_cols_j = {}
    for vName in in_cols :
         in_cols_j[vName] = cfg.getint("cols_in",  vName)
    #brands,   discount     = config_read(cfgFName, 'discount')
    #for k in discount.keys():
    #    discount[k] = (100 - int(discount[k]))/100
    #print(discount)

    outFileRUR1 = open( csvFNameRUR1, 'w', newline='', encoding='CP1251', errors='replace')
    outFileRUR2 = open( csvFNameRUR2, 'w', newline='', encoding='CP1251', errors='replace')
    outFileUSD1 = open( csvFNameUSD1, 'w', newline='', encoding='CP1251', errors='replace')
    outFileUSD2 = open( csvFNameUSD2, 'w', newline='', encoding='CP1251', errors='replace')
    outFileEUR1 = open( csvFNameEUR1, 'w', newline='', encoding='CP1251', errors='replace')
    outFileEUR2 = open( csvFNameEUR2, 'w', newline='', encoding='CP1251', errors='replace')
    csvWriterRUR1 = csv.DictWriter(outFileRUR1, fieldnames=out_cols )
    csvWriterRUR2 = csv.DictWriter(outFileRUR2, fieldnames=out_cols )
    csvWriterEUR1 = csv.DictWriter(outFileEUR1, fieldnames=out_cols )
    csvWriterEUR2 = csv.DictWriter(outFileEUR2, fieldnames=out_cols )
    csvWriterUSD1 = csv.DictWriter(outFileUSD1, fieldnames=out_cols )
    csvWriterUSD2 = csv.DictWriter(outFileUSD2, fieldnames=out_cols )
    csvWriterRUR1.writeheader()
    csvWriterRUR2.writeheader()
    csvWriterEUR1.writeheader()
    csvWriterEUR2.writeheader()
    csvWriterUSD1.writeheader()
    csvWriterUSD2.writeheader()

    '''                                     # Блок проверки свойств для распознавания групп      XLSX                                  
    for i in range(2393, 2397):                                                         
        i_last = i
        ccc = sheet.cell( row=i, column=in_cols_j['группа'] )
        print(i, ccc.value)
        print(ccc.font.name, ccc.font.sz, ccc.font.b, ccc.font.i, ccc.font.color.rgb, '------', ccc.fill.fgColor.rgb)
        print('------')
    '''
    '''                                     # Блок проверки свойств для распознавания групп      XLS                                  
    for i in range(0, 75):                                                         
        xfx = sheet.cell_xf_index(i, 0)
        xf  = book.xf_list[xfx]
        bgci  = xf.background.pattern_colour_index
        fonti = xf.font_index
        ccc = sheet.cell(i, 0)
        if ccc.value == None :
            print (i, colSGrp, 'Пусто!!!')
            continue
                                         # Атрибуты шрифта для настройки конфига
        font = book.font_list[fonti]
        print( '---------------------- Строка', i, '-----------------------', sheet.cell(i, 0).value)
        print( 'background_colour_index=',bgci)
        print( 'fonti=', fonti, '           xf.alignment.indent_level=', xf.alignment.indent_level)
        print( 'bold=', font.bold)
        print( 'weight=', font.weight)
        print( 'height=', font.height)
        print( 'italic=', font.italic)
        print( 'colour_index=', font.colour_index )
        print( 'name=', font.name)
    return
    '''

    recOut  ={}
#   for i in range(1, sheet.nrows) :                                     # xls
    for i in range(1, sheet.max_row +1):                                 # xlsx
        i_last = i
        try:
            impValues = getXlsxString(sheet, i, in_cols_j)
            #impValues = getXlsString(sheet, i, in_cols_j)
            #print( impValues['закупка'])
            if impValues['закупка']=='0': # (ccc.value == None) or (ccc2.value == None) :    # Пустая строка
                pass
            else:                                                        # Обычная строка
                if impValues['закупка'] == '0.1':
                    impValues['валюта1'] = 'USD'
                if impValues['продажа'] == '0.1':
                    impValues['валюта2'] = 'USD'

                for outColName in out_template.keys() :
                    shablon = out_template[outColName]
                    for key in impValues.keys():
                        if shablon.find(key) >= 0 :
                            shablon = shablon.replace(key, impValues[key])
                    if (outColName == 'закупка') and ('*' in shablon) :
                        vvvv = float( shablon[ :shablon.find('*')     ] )
                        shablon = str( float(vvvv) * brand_koeft )
                    recOut[outColName] = shablon

                if recOut['валюта1']=='USD':
                    csvWriterUSD1.writerow(recOut)
                elif recOut['валюта1']=='EUR':
                    csvWriterEUR1.writerow(recOut)
                elif recOut['валюта1']=='руб.':
                    csvWriterRUR1.writerow(recOut)
                else :
                    log.error('Не распознана валюта1 "%s" ')

                if recOut['валюта2']=='USD':
                    csvWriterUSD2.writerow(recOut)
                elif recOut['валюта2']=='EUR':
                    csvWriterEUR2.writerow(recOut)
                elif recOut['валюта2']=='руб.':
                    csvWriterRUR2.writerow(recOut)
                else :
                    log.error('Не распознана валюта2 "%s" ')

        except Exception as e:
            print(e)
            if str(e) == "'NoneType' object has no attribute 'rgb'":
                pass
            else:
                log.debug('Exception: <' + str(e) + '> при обработке строки ' + str(i) +'.' )

    log.info('Обработано ' +str(i_last)+ ' строк.')
    outFileRUR1.close()
    outFileUSD1.close()
    outFileEUR1.close()
    outFileRUR2.close()
    outFileUSD2.close()
    outFileEUR2.close()



def download( cfg ):
    from selenium import webdriver
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.remote.remote_connection import LOGGER
    LOGGER.setLevel(logging.WARNING)
    
    retCode     = False
    filename_new= cfg.get('download','filename_new')
    filename_old= cfg.get('download','filename_old')
    login       = cfg.get('download','login'    )
    password    = cfg.get('download','password' )
    url_lk      = cfg.get('download','url_lk'   )
    url_file    = cfg.get('download','url_file' )

    download_path= os.path.join(os.getcwd(), 'tmp')
    if not os.path.exists(download_path):
        os.mkdir(download_path)

    for fName in os.listdir(download_path) :
        os.remove( os.path.join(download_path, fName))
    dir_befo_download = set(os.listdir(download_path))
        
    if os.path.exists('geckodriver.log') : os.remove('geckodriver.log')
    try:
        ffprofile = webdriver.FirefoxProfile()
        ffprofile.set_preference("browser.download.dir", download_path)
        ffprofile.set_preference("browser.download.folderList",2);
        ffprofile.set_preference("browser.helperApps.neverAsk.saveToDisk", 
                ",application/octet-stream" + 
                ",application/vnd.ms-excel" + 
                ",application/vnd.msexcel" + 
                ",application/x-excel" + 
                ",application/x-msexcel" + 
                ",application/zip" + 
                ",application/xls" + 
                ",application/vnd.ms-excel" +
                ",application/vnd.ms-excel.addin.macroenabled.12" +
                ",application/vnd.ms-excel.sheet.macroenabled.12" +
                ",application/vnd.ms-excel.template.macroenabled.12" +
                ",application/vnd.ms-excelsheet.binary.macroenabled.12" +
                ",application/vnd.ms-fontobject" +
                ",application/vnd.ms-htmlhelp" +
                ",application/vnd.ms-ims" +
                ",application/vnd.ms-lrm" +
                ",application/vnd.ms-officetheme" +
                ",application/vnd.ms-pki.seccat" +
                ",application/vnd.ms-pki.stl" +
                ",application/vnd.ms-word.document.macroenabled.12" +
                ",application/vnd.ms-word.template.macroenabed.12" +
                ",application/vnd.ms-works" +
                ",application/vnd.ms-wpl" +
                ",application/vnd.ms-xpsdocument" +
                ",application/vnd.openofficeorg.extension" +
                ",application/vnd.openxmformats-officedocument.wordprocessingml.document" +
                ",application/vnd.openxmlformats-officedocument.presentationml.presentation" +
                ",application/vnd.openxmlformats-officedocument.presentationml.slide" +
                ",application/vnd.openxmlformats-officedocument.presentationml.slideshw" +
                ",application/vnd.openxmlformats-officedocument.presentationml.template" +
                ",application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" +
                ",application/vnd.openxmlformats-officedocument.spreadsheetml.template" +
                ",application/vnd.openxmlformats-officedocument.wordprocessingml.template" +
                ",application/x-ms-application" +
                ",application/x-ms-wmd" +
                ",application/x-ms-wmz" +
                ",application/x-ms-xbap" +
                ",application/x-msaccess" +
                ",application/x-msbinder" +
                ",application/x-mscardfile" +
                ",application/x-msclip" +
                ",application/x-msdownload" +
                ",application/x-msmediaview" +
                ",application/x-msmetafile" +
                ",application/x-mspublisher" +
                ",application/x-msschedule" +
                ",application/x-msterminal" +
                ",application/x-mswrite" +
                ",application/xml" +
                ",application/xml-dtd" +
                ",application/xop+xml" +
                ",application/xslt+xml" +
                ",application/xspf+xml" +
                ",application/xv+xml" +
                ",application/excel")
        if os.name == 'posix':
            driver = webdriver.Firefox(ffprofile, executable_path=r'geckodriver')  # , executable_path=r'/usr/local/bin/geckodriver')
        elif os.name == 'nt':
            driver = webdriver.Firefox(ffprofile)
        driver.implicitly_wait(10)
        driver.set_page_load_timeout(10)

        driver.get(url_lk)
        time.sleep(1)
        driver.find_element_by_name("USER_LOGIN").clear()
        driver.find_element_by_name("USER_LOGIN").send_keys(login)
        driver.find_element_by_name("USER_PASSWORD").clear()
        driver.find_element_by_name("USER_PASSWORD").send_keys(password)
        driver.find_element_by_name("Login").click()
        time.sleep(1)
        try:
            time.sleep(1)
            driver.get(url_file)
            time.sleep(10)
        except Exception as e:
            log.debug(e)
        #print(driver.page_source)
        #driver.find_element_by_css_selector("input.button-container-m.btn_ExportAll").click()
        #time.sleep(50)

    except Exception as e:
        log.debug('Exception: <' + str(e) + '>')

    driver.quit()
    dir_afte_download = set(os.listdir(download_path))
    new_files = list( dir_afte_download.difference(dir_befo_download))
    print(new_files)
    if len(new_files) == 0 :        
        log.error( 'Не удалось скачать файл прайса ')
        return False
    elif len(new_files)>1 :
        log.error( 'Скачалось несколько файлов. Надо разбираться ...')
        return False
    else:   
        new_file = new_files[0]                                                     # загружен ровно один файл. 
        new_ext  = os.path.splitext(new_file)[-1].lower()
        DnewFile = os.path.join( download_path,new_file)
        new_file_date = os.path.getmtime(DnewFile)
        log.info( 'Скачанный файл ' +new_file + ' имеет дату ' + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(new_file_date) ) )
        
        print(new_ext)
        if new_ext in ('.xls','.xlsx','.xlsb','.xlsm','.csv'):
            if os.path.exists( filename_new) and os.path.exists( filename_old): 
                os.remove( filename_old)
                os.rename( filename_new, filename_old)
            if os.path.exists( filename_new) :
                os.rename( filename_new, filename_old)
            shutil.copy2( DnewFile, filename_new)
            return True



def config_read( cfgFName ):
    cfg = configparser.ConfigParser(inline_comment_prefixes=('#'))
    if  os.path.exists('private.cfg'):     
        cfg.read('private.cfg', encoding='utf-8')
    if  os.path.exists(cfgFName):     
        cfg.read( cfgFName, encoding='utf-8')
    else: 
        log.debug('Нет файла конфигурации '+cfgFName)
    return cfg



def is_file_fresh(fileName, qty_days):
    qty_seconds = qty_days *24*60*60 
    if os.path.exists( fileName):
        price_datetime = os.path.getmtime(fileName)
    else:
        log.error('Не найден файл  '+ fileName)
        return False

    if price_datetime+qty_seconds < time.time() :
        file_age = round((time.time()-price_datetime)/24/60/60)
        log.error('Файл "'+fileName+'" устарел!  Допустимый период '+ str(qty_days)+' дней, а ему ' + str(file_age) )
        return False
    else:
        return True



def make_loger():
    global log
    logging.config.fileConfig('logging.cfg')
    log = logging.getLogger('logFile')



def processing(cfgFName):
    log.info('----------------------- Processing '+cfgFName )
    cfg = config_read(cfgFName)
    csvFNameRUR  = cfg.get('basic','filename_out_RUR')
    csvFNameEUR  = cfg.get('basic','filename_out_EUR')
    csvFNameUSD  = cfg.get('basic','filename_out_USD')
    priceFName= cfg.get('basic','filename_in')
    
    if cfg.has_section('download'):
        result = download(cfg)
    if is_file_fresh( priceFName, int(cfg.get('basic','срок годности'))):
        #os.system( dealerName + '_converter_xlsx.xlsm')
        #convert_csv2csv(cfg)
        convert_excel2csv(cfg)
    


def main( dealerName):
    """ Обработка прайсов выполняется согласно файлов конфигурации.
    Для этого в текущей папке должны быть файлы конфигурации, описывающие
    свойства файла и правила обработки. По одному конфигу на каждый 
    прайс или раздел прайса со своими правилами обработки
    """
    make_loger()
    log.info('          '+dealerName )
    for cfgFName in os.listdir("."):
        if cfgFName.startswith("cfg") and cfgFName.endswith(".cfg"):
            processing(cfgFName)


if __name__ == '__main__':
    myName = os.path.basename(os.path.splitext(sys.argv[0])[0])
    mydir    = os.path.dirname (sys.argv[0])
    print(mydir, myName)
    main( myName)
